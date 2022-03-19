from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook

def code_crawling() :
	driver = webdriver.Chrome('C:\dev\chromedriver_win32\chromedriver.exe')

	driver.get('http://all.jbnu.ac.kr/jbnu/sugang/sbjt/sbjt.html')
	driver.maximize_window()
	time.sleep(4)

	# 연도 선택
	for a in range(0):
		driver.find_element_by_id('mainframe_childframe_form_divSearch_spnYy_spindownbutton').click()
		time.sleep(2)

	# 조회 버튼 클릭
	search_bnt = driver.find_element_by_id('mainframe_childframe_form_divSearch_btnSearch').click()
	time.sleep(6)

	class_code = []
	cc = 0
	while True:
		for index in range(2, 26) :
			try :
		        # 강의 계획서 클릭
				find = driver.find_element_by_xpath('/html/body/div[1]/div/div/div/div[1]/div/div[5]/div[1]/div[4]/div[1]/div/div[%d]/div/div/div[23]/div/div'%index)
				find.click()
				time.sleep(1)
				# 과목코드 가져오기        
				code = driver.find_element_by_id('ubiviewer_%d_CELL4191_1_14'%cc).text
			except Exception: 
				try :
					# 과목코드 가져오기
					code = driver.find_element_by_id('ubiviewer_%d_CELL983_1_14'%cc).text
					time.sleep(1)
				except Exception:
					try :
						# 과목코드 가져오기
						code = driver.find_element_by_id('ubiviewer_%d_CELL3949_1_10'%cc).text
						time.sleep(1)
					except Exception:
						try :
							# 과목코드 가져오기
							code = driver.find_element_by_id('ubiviewer_%d_CELL983_1_29'%cc).text
							time.sleep(1)
						except Exception :
							try :
								# 과목코드 가져오기
								code = driver.find_element_by_id('ubiviewer_%d_CELL983_1_31'%cc).text
								time.sleep(1)
							except Exception as error : #강의계획서 없거나 찾지못한 예외 경우
								print(error)
								code = "NO"
								cc -= 1
								time.sleep(1)
			try :
				# 강의 계획서 닫기
				driver.find_element_by_id('mainframe_childframe_uniLessSubjtTblInqP01_titlebar_closebutton').click()
				time.sleep(1)
			except :
				#닫을게 없는경우
				print("hello")
			class_code += code
			cc += 1
			print(cc)
			print(code)
		for i in range(21) :
			driver.find_element_by_css_selector('#mainframe_childframe_form_grdLessSubjtExcel_vscrollbar_incbutton').click()
		time.sleep(3)
	return class_code
    
def list2exel(crawlingList) :
	wb = Workbook()
	ws = wb.create_sheet('Sheet1')
	ws = wb.active
	index = 2
	for code in crawlingList :
		ws.cell(index, 1, code)
		index += 1
	wb.save("C:/Users/ekffk/OneDrive/바탕 화면/class.xlsx")
    
list2exel(code_crawling())